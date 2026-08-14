#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Dingol 顶一 dingsel 选型数据提取脚本
只提取 4 张主表:4P 3PH / 4P 1PH / 2P 3PH / 2P 1PH
跳过:5 张变体工况表(40SB/27SB/150K/163K/1Hour) + 3 张 Marine
处理约定:
  - W11/W12 合成一个电压档(winding="11/12")
  - W13 / W14 是独立 380-416V 档,分开
  - winding_label / winding_conn 按 DINGOL 源表头如实写
  - 无数据字段省略(不写 0)
"""
import openpyxl, json, os, re

SRC='/Users/bob/.openclaw/workspace/media/inbound/openclaw-staged-7bf45d74-8e2c-40e8-b7bc-14be483593ff/Dingol_rating_table-rev13-20240222---f5c1c521-e129-4169-8cc7-d47a2886f20a.xlsx'
OUT='/Users/bob/.openclaw/workspace/genset-selector/dingol.json'

wb=openpyxl.load_workbook(SRC, data_only=True)

def fval(x):
    """取数值;None 或 'N/A' 或 '-' 返回 None"""
    if isinstance(x,(int,float)):
        return round(float(x), 4)
    if x is None: return None
    s=str(x).strip()
    if s in ('','-','N/A','n/a','NA'): return None
    try: return round(float(s),4)
    except: return None

# ============ 4P 3PH ============
# 表头(共用结构,板块1=行4-12 绕组11;板块2=行58-65 绕组11/12)
# 电压档定义: (code, label, conn, 绕组)
# 50Hz 6档: col2(380) col4(400) col6(415) col8(440) col10(380-416@W13) col12(380-416@W14)
# 60Hz 4档: col14(416) col16(440) col18(460) col20(480)
def extract_4p3ph(ws, title):
    recs=[]
    # ---- 板块1: 表头行4-12,数据行13-54 (model 行),实际找以 DG 开头
    # 板块1 电压档 (code, freq, pf, label, conn, winding)
    defs1=[
        ('380V','50Hz',0.8,'Star 380 / Delta 220','Star/Delta','11'),
        ('400V','50Hz',0.8,'Star 400 / Delta 230','Star/Delta','11'),
        ('415V','50Hz',0.8,'Star 415 / Delta 240','Star/Delta','11'),
        ('440V','50Hz',0.8,'Star 440 / Delta 254','Star/Delta','11'),
        ('380-416V(W13)','50Hz',0.8,'Star(W12&W13) 380-416 / Delta 220-240','Star(W12&W13)/Delta','13'),
        ('380-416V(W14)','50Hz',0.8,'Series/Parallel Star(W11&W14) 380-416 / 190-208','Series Star(W11&W14)/Parallel Star(W11&W14)','14'),
        ('416V','60Hz',0.8,'Star 416 / Delta 240','Star/Delta','11'),
        ('440V','60Hz',0.8,'Star 440 / Delta 254','Star/Delta','11'),
        ('460V','60Hz',0.8,'Star 460 / Delta 266','Star/Delta','11'),
        ('480V','60Hz',0.8,'Star 480 / Delta 277','Star/Delta','11'),
    ]
    # 每档 KVA/KW 列号 (1-index)
    cols=[(2,3),(4,5),(6,7),(8,9),(10,11),(12,13),(14,15),(16,17),(18,19),(20,21)]
    # 板块1 数据范围
    for r in range(13,56):
        m=ws.cell(row=r,column=1).value
        if m is None: continue
        m=str(m).strip()
        if not m or m.startswith('Model') or m.startswith('Power') or m.startswith('Three') \
           or m.startswith('Winding') or m.startswith('Rating') or m.startswith('Star') \
           or m.startswith('Delta') or m.startswith('Series') or m.startswith('Parallel') \
           or m.startswith('Volts') or m.startswith('Notes') or m.startswith('1.') \
           or m.startswith('2.') or m.startswith('3.') or m.startswith('4.'):
            continue
        if not re.match(r'^DG', m): continue
        for i,dc in enumerate(defs1):
            kwv=fval(ws.cell(row=r,column=cols[i][0]).value)
            kww=fval(ws.cell(row=r,column=cols[i][1]).value)
            if kwv is None and kww is None:
                continue  # 无数据省略
            rec={'type':'generator','brand':'dingol','model':m.strip(),
                 'frequency':dc[1],'winding_code':dc[0],'winding_label':dc[3],
                 'winding_conn':dc[4],'winding_no':dc[5],'wires':'12',
                 'pf':dc[2]}
            if kwv is not None:
                rec['cont_h_kVA']=kwv
                if kww is not None: rec['cont_h_kW']=kww
            elif kww is not None:
                # 理论上 KVA 必有,KW 可能缺;若只有 KW 也记
                rec['cont_h_kW']=kww
            recs.append(rec)
    # ---- 板块2: 表头行58-65,数据行66-79
    defs2=[
        ('380V','50Hz',0.8,'Star 380 / Delta 220','Star/Delta','11/12'),
        ('400V','50Hz',0.8,'Star 400 / Delta 230','Star/Delta','11/12'),
        ('415V','50Hz',0.8,'Star 415 / Delta 240','Star/Delta','11/12'),
        ('440V','50Hz',0.8,'Star 440 / Delta 254','Star/Delta','11/12'),
        ('380-416V(W13)','50Hz',0.8,'Star(W12&W13) 380-416 / Delta 220-240','Star(W12&W13)/Delta','13'),
        ('380-416V(W14)','50Hz',0.8,'Series/Parallel Star(W11&W14) 380-416 / 190-208','Series Star(W11&W14)/Parallel Star(W11&W14)','14'),
        ('416V','60Hz',0.8,'Star 416 / Delta 240','Star/Delta','11/12'),
        ('440V','60Hz',0.8,'Star 440 / Delta 254','Star/Delta','11/12'),
        ('460V','60Hz',0.8,'Star 460 / Delta 266','Star/Delta','11/12'),
        ('480V','60Hz',0.8,'Star 480 / Delta 277','Star/Delta','11/12'),
    ]
    for r in range(66,80):
        m=ws.cell(row=r,column=1).value
        if m is None: continue
        m=str(m).strip()
        if not m or not re.match(r'^DG', m): continue
        for i,dc in enumerate(defs2):
            kwv=fval(ws.cell(row=r,column=cols[i][0]).value)
            kww=fval(ws.cell(row=r,column=cols[i][1]).value)
            if kwv is None and kww is None: continue
            rec={'type':'generator','brand':'dingol','model':m.strip(),
                 'frequency':dc[1],'winding_code':dc[0],'winding_label':dc[3],
                 'winding_conn':dc[4],'winding_no':dc[5],'wires':'12',
                 'pf':dc[2]}
            if kwv is not None:
                rec['cont_h_kVA']=kwv
                if kww is not None: rec['cont_h_kW']=kww
            elif kww is not None: rec['cont_h_kW']=kww
            recs.append(rec)
    # 特例覆盖: DG634C 无 W11 -> 板块2 已标 11/12,需特殊? 用户层面标注在元数据,数据层保留
    return recs

# ============ 单相(4P1PH / 2P1PH) ============
# 列块: 50Hz(PF0.8:220/230/240; PF1:220/230/240) 60Hz同
# 列号: 1=model; 每档 KVA/KW 两列. 块首列: 2,4,6(0.8) 8,10,12(1.0) [50Hz]
#       14,16,18(0.8) 20,22,24(1.0) [60Hz]
def extract_1ph(ws, freq50, freq60, w50, w60):
    recs=[]
    blocks=[]
    for pf in (0.8,1.0):
        for v in (220,230,240):
            blocks.append((freq50,pf,v,w50))
    for pf in (0.8,1.0):
        for v in (220,230,240):
            blocks.append((freq60,pf,v,w60))
    # 每块 KVA/KW 列: 块i(从0),首列=2+2*i
    for r in range(9, ws.max_row+1):
        m=ws.cell(row=r,column=1).value
        if m is None: continue
        m=str(m).strip()
        if not m or m.startswith('Notes') or m.startswith('1.') or m.startswith('Model'):
            continue
        if not re.match(r'^DG', m): continue
        for i,(freq,pf,v,w) in enumerate(blocks):
            c0=2+2*i
            kwv=fval(ws.cell(row=r,column=c0).value)
            kww=fval(ws.cell(row=r,column=c0+1).value)
            if kwv is None and kww is None: continue
            code=f'1ph-{v}V-{str(pf).replace(".","")}'
            label=f'1ph Series {v}V / Parallel {v//2}V (PF {pf})'
            rec={'type':'generator','brand':'dingol','model':m.strip(),
                 'frequency':freq,'winding_code':code,'winding_label':label,
                 'winding_conn':'Series/Parallel','winding_no':w,'wires':'12',
                 'pf':pf}
            if kwv is not None:
                rec['cont_h_kVA']=kwv
                if kww is not None: rec['cont_h_kW']=kww
            elif kww is not None: rec['cont_h_kW']=kww
            recs.append(rec)
    return recs

# ============ 2P 3PH ============
# 电压档: 50Hz(380/400/415/440) 60Hz(416/440/460/480), 绕组11
def extract_2p3ph(ws):
    recs=[]
    defs=[
        ('380V','50Hz',0.8,'Series Star 380 / Parallel Star 190 / Series Delta 220','Series Star/Parallel Star/Series Delta','11'),
        ('400V','50Hz',0.8,'Series Star 400 / Parallel Star 200 / Series Delta 230','Series Star/Parallel Star/Series Delta','11'),
        ('415V','50Hz',0.8,'Series Star 415 / Parallel Star 208 / Series Delta 240','Series Star/Parallel Star/Series Delta','11'),
        ('440V','50Hz',0.8,'Series Star 440 / Parallel Star 220 / Series Delta 254','Series Star/Parallel Star/Series Delta','11'),
        ('416V','60Hz',0.8,'Series Star 416 / Parallel Star 208 / Series Delta 240','Series Star/Parallel Star/Series Delta','11'),
        ('440V','60Hz',0.8,'Series Star 440 / Parallel Star 220 / Series Delta 254','Series Star/Parallel Star/Series Delta','11'),
        ('460V','60Hz',0.8,'Series Star 460 / Parallel Star 230 / Series Delta 266','Series Star/Parallel Star/Series Delta','11'),
        ('480V','60Hz',0.8,'Series Star 480 / Parallel Star 240 / Series Delta 277','Series Star/Parallel Star/Series Delta','11'),
    ]
    cols=[(2,3),(4,5),(6,7),(8,9),(10,11),(12,13),(14,15),(16,17)]
    for r in range(10, ws.max_row+1):
        m=ws.cell(row=r,column=1).value
        if m is None: continue
        m=str(m).strip()
        if not m or m.startswith('Notes') or m.startswith('1.'): continue
        if not re.match(r'^DG', m): continue
        for i,dc in enumerate(defs):
            kwv=fval(ws.cell(row=r,column=cols[i][0]).value)
            kww=fval(ws.cell(row=r,column=cols[i][1]).value)
            if kwv is None and kww is None: continue
            rec={'type':'generator','brand':'dingol','model':m.strip(),
                 'frequency':dc[1],'winding_code':dc[0],'winding_label':dc[3],
                 'winding_conn':dc[4],'winding_no':dc[5],'wires':'12','pf':dc[2]}
            if kwv is not None:
                rec['cont_h_kVA']=kwv
                if kww is not None: rec['cont_h_kW']=kww
            elif kww is not None: rec['cont_h_kW']=kww
            recs.append(rec)
    return recs


# ============ 备用表数据回填 (40SB->stdby_40c, 27SB->stdby_27c) ============
DEFS_V=('380V','400V','415V','440V','380-416V(W13)','380-416V(W14)','416V','440V','460V','480V')
DEFS_F=('50Hz','50Hz','50Hz','50Hz','50Hz','50Hz','60Hz','60Hz','60Hz','60Hz')
VCOLS=[2,4,6,8,10,12,14,16,18,20]
VKWC=[3,5,7,9,11,13,15,17,19,21]

def read_sb_sheet(ws):
    """读一张备用表(40SB/27SB)的 板块1+板块2,返回 {model:{(code,freq):(kva,kw)}}"""
    rate_rows=[r for r in range(1,ws.max_row+1) if ws.cell(row=r,column=1).value and str(ws.cell(row=r,column=1).value).strip()=='Rating']
    if len(rate_rows)<2:
        # 单板块
        first_dg=min((r for r in range(1,ws.max_row+1) if ws.cell(row=r,column=1).value and str(ws.cell(row=r,column=1).value).strip().startswith('DG')), default=None)
        ranges=[(first_dg, ws.max_row)]
    else:
        first_dg=min((r for r in range(1,ws.max_row+1) if ws.cell(row=r,column=1).value and str(ws.cell(row=r,column=1).value).strip().startswith('DG')), default=None)
        ranges=[(first_dg, rate_rows[1]-2), (rate_rows[1]+1, ws.max_row)]
    out={}
    for (r0,r1) in ranges:
        if r0 is None: continue
        for r in range(r0, r1+1):
            m=ws.cell(row=r,column=1).value
            if m is None: continue
            m=str(m).strip()
            if not m.startswith('DG'): continue
            rec={}
            for i in range(10):
                kva=fval(ws.cell(row=r,column=VCOLS[i]).value)
                kw=fval(ws.cell(row=r,column=VKWC[i]).value)
                if kva is not None or kw is not None:
                    rec[(DEFS_V[i],DEFS_F[i])]=(kva,kw)
            out[m]=rec
    return out

def merge_sb(recs, sb_data, code_key, kw_key):
    """把备用数据回填进 main 记录。
    recs: 主记录列表; sb_data: {model:{(code,freq):(kva,kw)}};
    code_key/kw_key: 字段前缀如 'stdby_40c_kVA'/'stdby_40c_kW'
    只回填主记录中已存在的 (model,freq,code) 档。"""
    bykey={}
    for x in recs:
        bykey[(x['model'],x['frequency'],x['winding_code'])]=x
    filled=0
    for m,sbrecs in sb_data.items():
        for (code,freq),(kva,kw) in sbrecs.items():
            k=(m,freq,code)
            x=bykey.get(k)
            if x is None:
                continue  # 主表无此档,跳过
            if kva is not None:
                x[code_key]=kva; filled+=1
            if kw is not None:
                x[kw_key]=kw
    return filled


if __name__=='__main__':
    all_recs=[]
    import sys as _sys
    _write = '--write' in _sys.argv
    stats={}
    ws=wb.worksheets[0]
    r=extract_4p3ph(ws,'4P 3PH')
    all_recs+=r; stats['4P 3PH']=len(r)
    ws=wb.worksheets[1]
    r=extract_1ph(ws,'50Hz','60Hz','05','06')
    all_recs+=r; stats['4P 1PH']=len(r)
    ws=wb.worksheets[2]
    r=extract_2p3ph(ws)
    all_recs+=r; stats['2P 3PH']=len(r)
    ws=wb.worksheets[3]
    r=extract_1ph(ws,'50Hz','60Hz','05','06')
    all_recs+=r; stats['2P 1PH']=len(r)
    print('=== 各表记录数 ===', stats, '总计', len(all_recs))
    # 校验1: 所有记录必有 cont_h_kVA
    no_kva=[x for x in all_recs if 'cont_h_kVA' not in x]
    print('校验: 无 cont_h_kVA 记录数:', len(no_kva))
    # 校验2: 重复键(型号+频率+绕组)检查
    seen={}
    dup=[]
    for x in all_recs:
        k=(x['model'],x['frequency'],x['winding_code'])
        if k in seen: dup.append(k)
        seen[k]=1
    print('校验: 重复(型号+频率+绕组) 键数:', len(dup), dup[:10])
    # 校验3: KW=KVA*pf 一致性(有偏差>0.5% 列出)
    bad=[]
    for x in all_recs:
        if 'cont_h_kW' in x and 'cont_h_kVA' in x and 'pf' in x:
            exp=x['cont_h_kVA']*x['pf']
            if exp and abs(x['cont_h_kW']-exp)/exp>0.005:
                bad.append((x['model'],x['frequency'],x['winding_code'],x['cont_h_kVA'],x['cont_h_kW'],round(exp,2)))
    print('校验: KW≠KVA*pf 偏差>0.5%% 记录数:', len(bad))
    for b in bad[:20]: print('   ', b)
    # 汇总模型数
    models=sorted({x['model'] for x in all_recs})
    print('型号总数:', len(models))
    sb405=read_sb_sheet(wb.worksheets[4])
    sb275=read_sb_sheet(wb.worksheets[5])
    n1=merge_sb(all_recs, sb405, 'stdby_40c_kVA','stdby_40c_kW')
    n2=merge_sb(all_recs, sb275, 'stdby_27c_kVA','stdby_27c_kW')
    print('回填 stdby_40c:', n1, '档  | stdby_27c:', n2, '档')
    if _write:
        with open(OUT,'w',encoding='utf-8') as f:
            json.dump(all_recs,f,ensure_ascii=False,indent=1)
        print('已写入', OUT, len(all_recs),'条')
    else:
        print(models)
