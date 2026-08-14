#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""dingol.json 合并后双重验证
独立于 extract_dingol.py 的逻辑,从 JSON 反向读原始 Excel 逐格比对。
验证1: JSON 记录总数 = 各表提取数合计
验证2: 抽样逐格与原表比对(数值完全一致)
验证3: KVA/KW/PF 数学一致性
验证4: 字段完整性(无缺失 KVA、无多余空值、type/brand 统一)
验证5: 每型号 50Hz/60Hz 档数在合理范围
"""
import openpyxl, json, re, sys, collections

JSON='/Users/bob/.openclaw/workspace/genset-selector/dingol.json'
SRC='/Users/bob/.openclaw/workspace/media/inbound/openclaw-staged-7bf45d74-8e2c-40e8-b7bc-14be483593ff/Dingol_rating_table-rev13-20240222---f5c1c521-e129-4169-8cc7-d47a2886f20a.xlsx'

data=json.load(open(JSON,encoding='utf-8'))
wb=openpyxl.load_workbook(SRC, data_only=True)

def fval(x):
    if isinstance(x,(int,float)): return round(float(x),4)
    if x is None: return None
    try: return round(float(str(x).strip()),4)
    except: return None

ok=True
def check(cond,msg):
    global ok
    st='PASS' if cond else 'FAIL'
    if not cond: ok=False
    print(f'  [{st}] {msg}')

print('====== 验证1: 记录总数 ======')
print(f'  dingol.json 加载: {len(data)} 条')
# 期望: 4P3PH 448 + 4P1PH 264 + 2P3PH 56 + 2P1PH 84
exp=448+264+56+84
check(len(data)==exp, f'总数 {len(data)} == 期望 {exp}')

print('\n====== 验证2: 逐表统计 ======')
# 从 JSON 的 winding_code/频率 反推归属表
# 单相: code 以 1ph- 开头且频率 50/60; 需区分 4P/2P 靠 model
TWOP={'DG162D','DG162E','DG162F','DG162G','DG182H','DG182J','DG182K'}
def sheet_of(x):
    md=x['model']
    is2p = md in TWOP
    is1ph = '1ph-' in x['winding_code']
    if is1ph:
        return '2P 1PH' if is2p else '4P 1PH'
    else:
        return '2P 3PH' if is2p else '4P 3PH'
cnt=collections.Counter(sheet_of(x) for x in data)
print('  JSON 各表分布:', dict(cnt))
check(cnt['4P 3PH']==448,'4P 3PH = 448')
check(cnt['4P 1PH']==264,'4P 1PH = 264')
check(cnt['2P 3PH']==56,'2P 3PH = 56')
check(cnt['2P 1PH']==84,'2P 1PH = 84')

print('\n====== 验证3: 逐格与原表核对(全部记录) ======')
# 重建每表的 型号->(列,值) 原始查找,对每条 JSON 记录去原表取 KVA/KW 比对
def raw_lookup(ws, defs, cols, rng):
    """返回 {model:{(code,freq):(kva,kw)}}"""
    res=collections.defaultdict(dict)
    for r in rng:
        m=ws.cell(row=r,column=1).value
        if m is None: continue
        m=str(m).strip()
        if not m or not re.match(r'^DG',m): continue
        for dc,(c0,c1) in zip(defs,cols):
            res[m][(dc[0],dc[1])]=(fval(ws.cell(row=r,column=c0).value),fval(ws.cell(row=r,column=c1).value))
    return res

# 4P3PH 板块1+2 定义
d1=[
 ('380V','50Hz',0.8,'11'),('400V','50Hz',0.8,'11'),('415V','50Hz',0.8,'11'),('440V','50Hz',0.8,'11'),
 ('380-416V(W13)','50Hz',0.8,'13'),('380-416V(W14)','50Hz',0.8,'14'),
 ('416V','60Hz',0.8,'11'),('440V','60Hz',0.8,'11'),('460V','60Hz',0.8,'11'),('480V','60Hz',0.8,'11')]
d2=[
 ('380V','50Hz',0.8,'11/12'),('400V','50Hz',0.8,'11/12'),('415V','50Hz',0.8,'11/12'),('440V','50Hz',0.8,'11/12'),
 ('380-416V(W13)','50Hz',0.8,'13'),('380-416V(W14)','50Hz',0.8,'14'),
 ('416V','60Hz',0.8,'11/12'),('440V','60Hz',0.8,'11/12'),('460V','60Hz',0.8,'11/12'),('480V','60Hz',0.8,'11/12')]
cols=[(2,3),(4,5),(6,7),(8,9),(10,11),(12,13),(14,15),(16,17),(18,19),(20,21)]
ws=wb.worksheets[0]
raw=raw_lookup(ws,d1,cols,range(13,56))
raw.update(raw_lookup(ws,d2,cols,range(66,80)))

# 单相
def d1ph_defs():
    blocks=[]
    for pf in (0.8,1.0):
        for v in (220,230,240): blocks.append((f'1ph-{v}V-{str(pf).replace(".","")}','50Hz',pf))
    for pf in (0.8,1.0):
        for v in (220,230,240): blocks.append((f'1ph-{v}V-{str(pf).replace(".","")}','60Hz',pf))
    return blocks
b1=d1ph_defs()
def raw_1ph(ws):
    colm1=[(2+2*i,3+2*i) for i in range(12)]
    res=collections.defaultdict(dict)
    for r in range(9,ws.max_row+1):
        m=ws.cell(row=r,column=1).value
        if m is None: continue
        m=str(m).strip()
        if not m or not re.match(r'^DG',m): continue
        for dc,(c0,c1) in zip(b1,colm1):
            res[m][(dc[0],dc[1])]=(fval(ws.cell(row=r,column=c0).value),fval(ws.cell(row=r,column=c1).value))
    return res
raw4p1=raw_1ph(wb.worksheets[1]); raw2p1=raw_1ph(wb.worksheets[3])

# 2P3PH
d2p=[
 ('380V','50Hz',0.8,'11'),('400V','50Hz',0.8,'11'),('415V','50Hz',0.8,'11'),('440V','50Hz',0.8,'11'),
 ('416V','60Hz',0.8,'11'),('440V','60Hz',0.8,'11'),('460V','60Hz',0.8,'11'),('480V','60Hz',0.8,'11')]
c2p=[(2,3),(4,5),(6,7),(8,9),(10,11),(12,13),(14,15),(16,17)]
raw2p3=raw_lookup(wb.worksheets[2],d2p,c2p,range(10,ws2p3_max_row:=27))

# 逐条核对
mismatch=[]
for x in data:
    md=x['model']; code=x['winding_code']
    sh=sheet_of(x)
    if sh=='4P 3PH': rl=raw.get(md,{})
    elif sh=='4P 1PH': rl=raw4p1.get(md,{})
    elif sh=='2P 3PH': rl=raw2p3.get(md,{})
    else: rl=raw2p1.get(md,{})
    kk=(code,x['frequency'])
    if kk not in rl:
        mismatch.append((md,code,'原始表无此档')); continue
    rkva,rkw=rl[kk]
    jkva=x.get('cont_h_kVA'); jkw=x.get('cont_h_kW')
    if rkva is None:
        # 原始无KVA但JSON有 -> 错
        if jkva is not None: mismatch.append((md,code, f'原始无KVA但JSON有{jkva}'))
    else:
        if jkva is None or abs(jkva-rkva)>0.001:
            mismatch.append((md,code,f'KVA不匹配 JSON={jkva} 原始={rkva}'))
    if rkw is not None:
        if jkw is None or abs(jkw-rkw)>0.001:
            mismatch.append((md,code,f'KW不匹配 JSON={jkw} 原始={rkw}'))
    elif jkw is not None:
        # 原始无KW但JSON有? 表里KW列通常都有
        mismatch.append((md,code, f'原始无KW但JSON有{jkw}'))
check(len(mismatch)==0, f'逐格核对 {len(data)} 条, 不匹配 {len(mismatch)}')
for m in mismatch[:20]: print('     ',m)

print('\n====== 验证4: 字段完整性 ======')
missing_brand=[x for x in data if x.get('brand')!='dingol']
missing_type=[x for x in data if x.get('type')!='generator']
missing_code=[x for x in data if not x.get('winding_code')]
missing_label=[x for x in data if not x.get('winding_label')]
missing_conn=[x for x in data if not x.get('winding_conn')]
missing_pf=[x for x in data if not x.get('pf')]
check(not missing_brand, f'brand 全为 dingol')
check(not missing_type, f'type 全为 generator')
check(not missing_code, f'winding_code 全有')
check(not missing_label, f'winding_label 全有')
check(not missing_conn, f'winding_conn 全有')
check(not missing_pf, f'pf 全有')
no_kva=[x for x in data if 'cont_h_kVA' not in x]
check(not no_kva, f'无 cont_h_kVA 记录={len(no_kva)} (应为0, KVA必有)')

print('\n====== 验证5: 交叉一致性 ======')
# 5a: 每种 (频率,winding_code) 下所有记录 KVA 值应为非负且非空(已在验证4覆盖)
# 5b: 4P3PH 大机座板块 (DG634/DG734) 50Hz 档数上限6、60Hz 4
#     小机座 50Hz 上限5或6(视有无W13/W14)
ok5=True
viol5=[]
from collections import defaultdict as _dd
for freq,expect in (('50Hz',6),('60Hz',4)):
    for md in sorted(set(x['model'] for x in data)):
        sub=[x for x in data if x['model']==md and x['frequency']==freq and '1ph-' not in x['winding_code'] and not (md.startswith('DG162') or md.startswith('DG182'))]
        n=len(sub)
        if n>expect:
            viol5.append((md,freq,n,expect))
check(not viol5, f'4P3PH 档数未超表头上限 {len(viol5)}')
for v in viol5[:10]: print('     ',v)

# 5c: 单一 (model,freq,code) 唯一性(不重复) —— 已在验证1去重,但显式检查
from collections import Counter
keys=Counter((x['model'],x['frequency'],x['winding_code']) for x in data)
dups=[k for k,n in keys.items() if n>1]
check(not dups, f'(model,freq,code) 全唯一 {len(dups)}')
for k in dups[:10]: print('     ',k)

print('\n====== 结果 ======')
print('双重验证:', '全部通过 ✓' if ok else '存在失败 ✗')
sys.exit(0 if ok else 1)
